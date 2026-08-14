#!/usr/bin/env python3
"""Generate a copy-ready JavaScript filler for the NTU TDR upload form.

Run this after building the final thesis PDF:

    pixi run python generate_tdr_upload_script.py

Metadata is collected from ntusetup.tex, main.tex, front/abstract.tex,
main.toc, and main.pdf. Values that cannot be inferred safely are prompted.
Use --help for plain-text overrides and non-interactive operation.
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import warnings
from datetime import date
from functools import lru_cache
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from urllib.parse import urljoin

import pymupdf
import requests
from openpyxl import load_workbook

warnings.filterwarnings(
    "ignore",
    message=r"Cannot parse header or footer.*",
    category=UserWarning,
    module=r"openpyxl\.worksheet\.header_footer",
)


class CollectionError(ValueError):
    pass


# TDR 的系所下拉選單偶爾與 ntusetup.tex 裡的所名不同（例如「物理學研究所」在
# TDR 上是「物理學系」）。遇到對不起來的情形時在這裡加一組對應即可，
# 或改用 --department 直接指定。
# TDR's department dropdown does not always use the same wording as
# ntusetup.tex. Add a mapping here when they disagree, or pass --department.
DEPARTMENT_ALIASES: dict[str, str] = {
    # "物理學研究所": "物理學系",
}
# ntusetup.tex 與 ntuthesis.cls 出廠時填的假資料。值還停在這些字串代表使用者
# 沒改過，此時改為詢問，以免把範例資料送進 TDR。
# The dummy values ntusetup.tex and ntuthesis.cls ship with. A field still
# holding one of these was never edited, so prompt instead of submitting it.
PLACEHOLDER_EMAIL = ("Email Address", "stitch@ntu.edu.tw")
PLACEHOLDER_ORCID = ("0000-0000-0000-0000",)

NTU_CALENDAR_URL = "https://www.aca.ntu.edu.tw/w/aca/calendar"
HTTP_TIMEOUT_SECONDS = 20

EMAIL_PATTERN = re.compile(r"[^@\s]+@[^@\s]+\.[^@\s]+")
ORCID_PATTERN = re.compile(r"\d{4}-\d{4}-\d{4}-\d{3}[\dX]")
# 上傳頁面的 checkename()：名、空白，再接姓，兩者都是拉丁字母。
# The upload page's own checkename(): a given name, whitespace, then the rest,
# both written in Latin letters extended through Latin Extended-B.
LATIN_LETTERS = r"A-Za-zÀ-ÖØ-öø-ɏ"
ENGLISH_NAME_PATTERN = re.compile(rf"[{LATIN_LETTERS}-]+\s+[{LATIN_LETTERS}\s.-]+")

# scripts/fetch_academic_units.py regenerates this from NTU's published list.
ACADEMIC_UNITS_FILE = "ntu-academic-units.tex"

# Official thesis-submission deadlines determine the graduation semester.
THESIS_DEADLINES = {
    # Special announcement: https://www.aca.ntu.edu.tw/w/aca/GAADNews_25102010004349933
    "114-1": date(2026, 2, 25),
    "114-2": date(2026, 8, 25),
    # NTU 115 academic calendar: https://www.aca.ntu.edu.tw/w/aca/calendar
    "115-1": date(2027, 2, 22),
}


class CalendarLinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self._href: str | None = None
        self._text: list[str] = []

    def handle_starttag(
        self, tag: str, attrs: list[tuple[str, str | None]]
    ) -> None:
        if tag.lower() != "a":
            return
        self._href = dict(attrs).get("href")
        self._text = []

    def handle_data(self, data: str) -> None:
        if self._href is not None:
            self._text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "a" and self._href is not None:
            self.links.append(("".join(self._text).strip(), self._href))
            self._href = None
            self._text = []


def strip_comments(text: str) -> str:
    lines = []
    for line in text.splitlines():
        match = re.search(r"(?<!\\)%", line)
        lines.append(line[: match.start()] if match else line)
    return "\n".join(lines)


def braced_group(text: str, opening: int) -> tuple[str, int]:
    if opening >= len(text) or text[opening] != "{":
        raise CollectionError("Expected an opening brace while parsing LaTeX")
    depth = 0
    escaped = False
    for index in range(opening, len(text)):
        char = text[index]
        if escaped:
            escaped = False
        elif char == "\\":
            escaped = True
        elif char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                return text[opening + 1 : index], index + 1
    raise CollectionError("Unbalanced braces while parsing LaTeX")


def latex_to_plain(text: str) -> str:
    text = strip_comments(text)
    text = re.sub(r"\\(?:begin|end)\{[^{}]+\}", "", text)
    text = re.sub(
        r"\\texorpdfstring\s*\{([^{}]*)\}\s*\{[^{}]*\}", r"\1", text
    )
    unwrap = re.compile(
        r"\\(?:textbf|textit|emph|mbox|textrm|textsf|texttt|mathrm|mathbf|"
        r"mathit|operatorname|url)\s*\{([^{}]*)\}"
    )
    previous = None
    while previous != text:
        previous = text
        text = unwrap.sub(r"\1", text)
    for source, target in {
        r"\&": "&",
        r"\%": "%",
        r"\#": "#",
        r"\_": "_",
        r"\$": "$",
        r"\{": "{",
        r"\}": "}",
        "~": " ",
    }.items():
        text = text.replace(source, target)
    text = re.sub(r"\\[(),\[\]]", "", text).replace("$", "")
    text = re.sub(r"\\[A-Za-z@]+\*?", "", text)
    return text.replace("{", "").replace("}", "")


def normalize_paragraphs(text: str) -> str:
    """Remove line wrapping inside paragraphs; preserve paragraph breaks."""
    text = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return ""
    paragraphs = re.split(r"\n\s*\n+", text)
    return "\n\n".join(
        re.sub(r"\s+", " ", paragraph).strip()
        for paragraph in paragraphs
        if paragraph.strip()
    )


def normalize_abstract(text: str) -> str:
    """Use one newline between paragraphs and none within a paragraph."""
    return normalize_paragraphs(text).replace("\n\n", "\n")


def parse_ntusetup(path: Path) -> dict[str, str]:
    text = strip_comments(path.read_text(encoding="utf-8"))
    marker = re.search(r"\\ntusetup\s*\{", text)
    if not marker:
        raise CollectionError(f"Could not find \\ntusetup in {path}")
    block, _ = braced_group(text, marker.end() - 1)
    values: dict[str, str] = {}
    cursor = 0
    pattern = re.compile(r"([A-Za-z][\w*-]*)\s*=\s*\{")
    while match := pattern.search(block, cursor):
        value, cursor = braced_group(block, match.end() - 1)
        values[match.group(1)] = latex_to_plain(value).strip()
    return values


COMMITTEE_REQUIRED = ("name", "name*", "email", "title")
COMMITTEE_TITLES = ("指導教授", "共同指導教授", "口試委員")


def parse_committee(path: Path) -> list[dict[str, str]]:
    """Read the \\ntucommittee entries from ntusetup.tex, in source order.

    The class defines \\ntucommittee as a no-op, so this list exists purely to
    fill in TDR's oral examination committee section.
    """
    text = strip_comments(path.read_text(encoding="utf-8"))
    field = re.compile(r"([A-Za-z][\w*-]*)\s*=\s*\{")
    members: list[dict[str, str]] = []

    for entry in re.finditer(r"\\ntucommittee\s*\{", text):
        block, _ = braced_group(text, entry.end() - 1)
        fields: dict[str, str] = {}
        cursor = 0
        while match := field.search(block, cursor):
            value, cursor = braced_group(block, match.end() - 1)
            fields[match.group(1)] = latex_to_plain(value).strip()

        position = len(members) + 1
        missing = [key for key in COMMITTEE_REQUIRED if not fields.get(key)]
        if missing:
            raise CollectionError(
                f"Committee member {position} in {path} is missing: "
                + ", ".join(missing)
            )
        if fields["title"] not in COMMITTEE_TITLES:
            raise CollectionError(
                f"Committee member {position} has title {fields['title']!r}; "
                f"expected one of {', '.join(COMMITTEE_TITLES)}"
            )
        # 表單自己會擋下格式錯誤的欄位，在這裡先擋，錯誤才不會等到最後一格才跳出。
        # The form rejects these itself; catching them here reports the entry at
        # fault instead of a dialog after the last field is typed.
        if not ENGLISH_NAME_PATTERN.fullmatch(fields["name*"]):
            raise CollectionError(
                f"Committee member {position} has name* {fields['name*']!r}; "
                "put the given name before the family name and use Latin "
                "letters only, as in Ming-Wen Li"
            )
        if not EMAIL_PATTERN.fullmatch(fields["email"]):
            raise CollectionError(
                f"Committee member {position} has email {fields['email']!r}, "
                "which is not an e-mail address"
            )
        if fields.get("ORCID") and not ORCID_PATTERN.fullmatch(fields["ORCID"]):
            raise CollectionError(
                f"Committee member {position} has ORCID {fields['ORCID']!r}; "
                "expected 0000-0000-0000-000X format"
            )

        members.append(
            {
                "nameZh": fields["name"],
                "nameEn": fields["name*"],
                "email": fields["email"],
                "title": fields["title"],
                "orcid": fields.get("ORCID", ""),
            }
        )

    if not members:
        raise CollectionError(f"No \\ntucommittee entries found in {path}")
    # 表單只在第一個委員區塊提供「指導教授」，因此指導教授必須排在第一位。
    # TDR offers 指導教授 only in its first committee block, so the advisor has
    # to lead the list for the blocks and the entries to line up.
    if members[0]["title"] != "指導教授":
        raise CollectionError(
            "The committee list must start with the 指導教授 entry; TDR offers "
            "that title only in its first committee block"
        )
    for position, member in enumerate(members[1:], start=2):
        if member["title"] == "指導教授":
            raise CollectionError(
                f"Committee member {position} is a second 指導教授; after the "
                "first entry only 共同指導教授 and 口試委員 are allowed"
            )
    return members


def parse_academic_units(path: Path) -> tuple[set[tuple[str, str]], set[tuple[str, str]]]:
    """Read the college pairs and the per-college unit names from the data file.

    Returns the set of (Chinese, English) college pairs and the set of
    (college, unit) pairs. Units are keyed by the college of their own
    language: NTU's two pages do not list identical units, so the generator
    deliberately does not pair them across languages.
    """
    # The file's own header documents the two macros by example, so the
    # comments have to go before the declarations are counted.
    text = strip_comments(path.read_text(encoding="utf-8"))
    colleges = {
        (m.group(1), m.group(2))
        for m in re.finditer(r"\\ntu@declarecollege\{([^}]*)\}\{([^}]*)\}", text)
    }
    units = {
        (m.group(1), m.group(2))
        for m in re.finditer(r"\\ntu@declareinstitute\{([^}]*)\}\{([^}]*)\}", text)
    }
    if not colleges or not units:
        raise CollectionError(f"{path} contains no academic units")
    return colleges, units


def check_academic_units(setup: dict[str, str], root: Path) -> None:
    """Reject a college or institute that is not one NTU publishes.

    ntuthesis.cls only warns about this so that a half-filled ntusetup.tex
    still produces a PDF. By the time an upload script is being generated the
    names are going onto a submission, so the same checks are fatal here.
    """
    path = root / ACADEMIC_UNITS_FILE
    if not path.is_file():
        print(
            f"Warning: {ACADEMIC_UNITS_FILE} is missing, so the college and "
            "institute names were not checked; run "
            "scripts/fetch_academic_units.py to restore it",
            file=sys.stderr,
        )
        return

    colleges, units = parse_academic_units(path)
    college_zh, college_en = setup.get("college", ""), setup.get("college*", "")
    if (college_zh, college_en) not in colleges:
        raise CollectionError(
            f"college {college_zh!r} and college* {college_en!r} are not a "
            f"college pair in NTU's academic units; see {ACADEMIC_UNITS_FILE}"
        )
    for college, key, label in (
        (college_zh, "institute", "institute"),
        (college_en, "institute*", "institute*"),
    ):
        name = setup.get(key, "")
        if (college, name) not in units:
            raise CollectionError(
                f"{label} {name!r} is not listed under {college!r} in NTU's "
                f"academic units; see {ACADEMIC_UNITS_FILE}"
            )


def parse_language(path: Path) -> str:
    text = strip_comments(path.read_text(encoding="utf-8"))
    document = re.search(
        r"\\documentclass\s*\[(.*?)\]\s*\{ntuthesis\}", text, re.DOTALL
    )
    if not document:
        raise CollectionError(f"Could not find ntuthesis options in {path}")
    language = re.search(r"\blanguage\s*=\s*(chinese|english)\b", document.group(1))
    if not language:
        raise CollectionError(f"Could not infer the thesis language from {path}")
    return "中文" if language.group(1) == "chinese" else "English"


def extract_abstracts(path: Path) -> tuple[str, str]:
    text = strip_comments(path.read_text(encoding="utf-8"))

    def environment(name: str) -> str:
        match = re.search(
            rf"\\begin\{{{re.escape(name)}\}}(.*?)\\end\{{{re.escape(name)}\}}",
            text,
            re.DOTALL,
        )
        return normalize_abstract(latex_to_plain(match.group(1))) if match else ""

    return environment("abstract"), environment("abstract*")


def command_groups(line: str, command: str) -> list[str]:
    match = re.match(rf"\s*\\{re.escape(command)}\s*", line)
    if not match:
        return []
    groups = []
    cursor = match.end()
    while cursor < len(line):
        while cursor < len(line) and line[cursor].isspace():
            cursor += 1
        if cursor >= len(line) or line[cursor] != "{":
            break
        value, cursor = braced_group(line, cursor)
        groups.append(value)
    return groups


def toc_title(value: str) -> str:
    number = ""
    match = re.search(r"\\numberline\s*\{", value)
    if match:
        number, end = braced_group(value, match.end() - 1)
        value = value[: match.start()] + value[end:]
    value = re.sub(r"\\(?:MakeUppercase|protect)\b", "", value)
    words = re.sub(r"\s+", " ", latex_to_plain(value)).strip()
    number = latex_to_plain(number).strip()
    return f"{number} {words}".strip()


def extract_toc(path: Path) -> str:
    entries = []
    indents = {
        "part": "",
        "chapter": "",
        "section": "  ",
        "subsection": "    ",
        "subsubsection": "      ",
    }
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        groups = command_groups(line, "contentsline")
        if len(groups) < 3 or groups[0] not in indents:
            continue
        level, title, page = groups[:3]
        entries.append(
            f"{indents[level]}{toc_title(title)} {latex_to_plain(page).strip()}".rstrip()
        )
    return "\n".join(entries)


def pdf_text(path: Path) -> str:
    pymupdf.TOOLS.mupdf_display_errors(False)
    try:
        with pymupdf.open(path) as document:
            pages = [
                page.get_text(
                    "text", sort=True, clip=pymupdf.INFINITE_RECT()
                )
                for page in document
            ]
    except (pymupdf.FileDataError, RuntimeError, ValueError) as error:
        raise CollectionError(f"Could not read {path}: {error}") from error
    return "\f".join(pages) + "\f"


def last_printed_page(text: str) -> str:
    pages = text.split("\f")
    if pages and not pages[-1].strip():
        pages.pop()  # the PDF extractor ends a document with a form feed.

    candidates: list[tuple[int, int]] = []
    for page_index, page in enumerate(pages):
        tail = [line.strip() for line in page.splitlines() if line.strip()][-8:]
        numbers = []
        for line in tail:
            footer = re.fullmatch(
                r"(\d+)(?:\s+doi:\S+)?", line, re.IGNORECASE
            )
            if footer:
                numbers.append(int(footer.group(1)))
        if numbers:
            candidates.append((page_index, numbers[-1]))
    if not candidates:
        return ""

    # The manual says unnumbered leaves after the final numbered page still
    # count. Front matter does not: its Roman-numbered leaves occur before it.
    last_index, printed_number = max(candidates, key=lambda item: item[1])
    return str(printed_number + len(pages) - last_index - 1)


def normalize_references(text: str) -> str:
    lines = [line.strip() for line in text.splitlines()]
    bracketed_marker = re.compile(r"^\[\d+\]\s*")
    plain_marker = re.compile(r"^\d+[.)]\s+")
    marker = (
        bracketed_marker
        if any(bracketed_marker.match(line) for line in lines)
        else plain_marker
    )
    if any(marker.match(line) for line in lines):
        entries: list[str] = []
        current: list[str] = []
        for line in lines:
            if not line:
                continue
            if marker.match(line) and current:
                entries.append(" ".join(current))
                current = [line]
            else:
                current.append(line)
        if current:
            entries.append(" ".join(current))
        return "\n".join(re.sub(r"\s+", " ", entry) for entry in entries)
    if re.search(r"\n\s*\n", text.strip()):
        return normalize_paragraphs(text).replace("\n\n", "\n")
    # Manual override files must already contain one complete entry per line.
    return "\n".join(re.sub(r"\s+", " ", line) for line in lines if line)


def references_from_pdf(text: str) -> str:
    pages = text.split("\f")
    heading = re.compile(r"^\s*(?:References|參考文獻)\s*$", re.IGNORECASE)
    start: tuple[int, int] | None = None
    for page_number, page in enumerate(pages):
        for line_number, line in enumerate(page.splitlines()):
            if heading.match(line):
                start = page_number, line_number + 1
                break
        if start:
            break
    if not start:
        return ""

    selected = []
    appendix = re.compile(r"^\s*(?:Appendix|附錄)(?:\s|$)", re.IGNORECASE)
    for page_number in range(start[0], len(pages)):
        lines = pages[page_number].splitlines()
        if page_number == start[0]:
            lines = lines[start[1] :]
        for line in lines:
            if heading.match(line):
                continue
            if appendix.match(line):
                return normalize_references("\n".join(selected))
            footer = re.fullmatch(
                r"\s*(?:\d+(?:\s+doi:\S+)?|doi:\S+)\s*",
                line,
                re.IGNORECASE,
            )
            if not footer:
                selected.append(line)
    return normalize_references("\n".join(selected))


def calendar_workbook_links(html: str) -> dict[int, str]:
    parser = CalendarLinkParser()
    parser.feed(html)
    links: dict[int, str] = {}
    for label, href in parser.links:
        academic_year = re.search(r"(\d{3})\s*學年度.*行事曆", label)
        if not academic_year or not href.lower().split("?", 1)[0].endswith(".xlsx"):
            continue
        links.setdefault(int(academic_year.group(1)), urljoin(NTU_CALENDAR_URL, href))
    if not links:
        raise CollectionError("The NTU calendar page contains no .xlsx calendars")
    return links


def calendar_event_date(academic_year: int, month: int, day: int) -> date:
    gregorian_year = academic_year + (1911 if month >= 8 else 1912)
    return date(gregorian_year, month, day)


def deadlines_from_calendar_workbook(
    payload: bytes, academic_year: int
) -> dict[str, date]:
    try:
        workbook = load_workbook(BytesIO(payload), data_only=True, read_only=True)
    except Exception as error:
        raise CollectionError(
            f"Could not read the NTU {academic_year} academic calendar: {error}"
        ) from error

    deadlines: dict[str, date] = {}
    try:
        for worksheet in workbook.worksheets:
            month: int | None = None
            for row in worksheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value is not None]
                for text in values:
                    if month_match := re.fullmatch(r"(\d{1,2})月", text):
                        month = int(month_match.group(1))
                if month is None:
                    continue

                for text in values:
                    day_match = re.match(r"^\s*(\d{1,2})日", text)
                    if not day_match:
                        continue
                    event_date = calendar_event_date(
                        academic_year, month, int(day_match.group(1))
                    )
                    if "學位論文" in text and "繳交截止" in text:
                        semester_match = re.search(r"第([一二])學期", text)
                        if not semester_match:
                            continue
                        if semester_match.group(1) == "一":
                            semester = f"{academic_year}-1"
                        else:
                            year = academic_year - 1 if month >= 8 else academic_year
                            semester = f"{year}-2"
                        deadlines[semester] = event_date
                    elif "第二學期" in text and "上課開始" in text:
                        deadlines.setdefault(f"{academic_year}-1", event_date)
    finally:
        workbook.close()
    return deadlines


@lru_cache(maxsize=1)
def official_thesis_deadlines() -> dict[str, date]:
    headers = {"User-Agent": "ntu-thesis-template TDR metadata generator"}
    try:
        response = requests.get(
            NTU_CALENDAR_URL, headers=headers, timeout=HTTP_TIMEOUT_SECONDS
        )
        response.raise_for_status()
        links = calendar_workbook_links(response.content.decode("utf-8"))
    except (requests.RequestException, UnicodeDecodeError) as error:
        raise CollectionError(f"Could not read the NTU calendar page: {error}") from error

    deadlines: dict[str, date] = {}
    failures: list[str] = []
    for academic_year, url in sorted(links.items()):
        try:
            response = requests.get(url, headers=headers, timeout=HTTP_TIMEOUT_SECONDS)
            response.raise_for_status()
            deadlines.update(
                deadlines_from_calendar_workbook(response.content, academic_year)
            )
        except (requests.RequestException, CollectionError) as error:
            failures.append(f"{academic_year}: {error}")
    if not deadlines:
        detail = "; ".join(failures) or "no deadline events were found"
        raise CollectionError(f"Could not obtain NTU thesis deadlines: {detail}")
    return deadlines


def semester_for_date(
    thesis_date: date, deadlines: dict[str, date]
) -> str | None:
    for semester, deadline in sorted(deadlines.items(), key=lambda item: item[1]):
        if thesis_date <= deadline:
            return semester
    return None


def graduation_semester(thesis_date: date) -> str:
    if semester := semester_for_date(thesis_date, THESIS_DEADLINES):
        return semester

    deadlines = official_thesis_deadlines().copy()
    if 3 <= thesis_date.month <= 8:
        semester = f"{thesis_date.year - 1912}-2"
        deadlines.setdefault(semester, date(thesis_date.year, 8, 25))
    deadlines.update(THESIS_DEADLINES)  # Special announcements have priority.
    if semester := semester_for_date(thesis_date, deadlines):
        return semester
    raise CollectionError(
        f"No published NTU thesis deadline covers {thesis_date.isoformat()}"
    )


def academic_period(iso_date: str) -> tuple[str, str, str]:
    """Calculate the graduation period from the thesis document date."""
    thesis_date = date.fromisoformat(iso_date)
    graduation_year, semester = graduation_semester(thesis_date).split("-")
    publication_year = thesis_date.year - 1911
    return graduation_year, semester, str(publication_year)


def read_text(path_value: str | None, root: Path) -> str:
    if not path_value:
        return ""
    path = Path(path_value)
    path = path if path.is_absolute() else root / path
    if not path.is_file():
        raise CollectionError(f"Text file does not exist: {path}")
    return path.read_text(encoding="utf-8")


def require_ignored_output(output: Path, root: Path) -> None:
    """Refuse to put submission/personal data in a tracked workspace path."""
    try:
        relative = output.resolve().relative_to(root.resolve())
    except ValueError:
        return
    result = subprocess.run(
        ["git", "check-ignore", "--quiet", "--", str(relative)],
        cwd=root,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        raise CollectionError(
            f"Output inside the thesis repository must be Git-ignored: {output}"
        )


def placeholder(value: str | None, *unset: str) -> str:
    """Treat a still-unedited ntusetup.tex placeholder as if it were missing.

    The template ships every key filled in with an obvious dummy so that the
    cover typesets out of the box. An untouched value therefore has to be
    caught here rather than silently submitted to TDR. Pass every sentinel the
    field could still be holding: the class default and the template's dummy.
    """
    text = (value or "").strip()
    return "" if text in unset else text


def choose(
    supplied: str | None,
    default: str,
    label: str,
    interactive: bool,
    required: bool = True,
) -> str:
    if supplied is not None:
        return supplied.strip()
    if default or not required:
        return default
    if interactive:
        while True:
            value = input(f"{label}: ").strip()
            if value:
                return value
            print("  This value is required.", file=sys.stderr)
    raise CollectionError(f"Missing {label}; pass the corresponding command-line option")


def validate(data: dict[str, object]) -> None:
    required = (
        "titleZh",
        "titleEn",
        "department",
        "graduationYear",
        "semester",
        "publicationYear",
        "authorZh",
        "authorEn",
        "email",
        "language",
        "defenseDate",
        "pages",
        "abstractZh",
        "abstractEn",
        "tableOfContents",
        "references",
    )
    missing = [key for key in required if not str(data.get(key, "")).strip()]
    if missing:
        raise CollectionError("Missing required field(s): " + ", ".join(missing))
    if not data.get("committee"):
        raise CollectionError("Missing the oral examination committee list")
    if data["semester"] not in {"1", "2"}:
        raise CollectionError("semester must be 1 or 2")
    for key in ("graduationYear", "publicationYear", "pages"):
        if not re.fullmatch(r"\d+", str(data[key])) or int(str(data[key])) <= 0:
            raise CollectionError(f"{key} must be a positive integer")
    try:
        date.fromisoformat(str(data["defenseDate"]))
    except ValueError as error:
        raise CollectionError("defenseDate must use YYYY-MM-DD") from error
    if data["language"] not in {"中文", "English"}:
        raise CollectionError('language must be "中文" or "English"')
    if "," in str(data["authorEn"]):
        raise CollectionError("authorEn must put the given name before the family name")
    email_pattern = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
    for key in ("email", "email2"):
        if data[key] and not email_pattern.fullmatch(str(data[key])):
            raise CollectionError(f"Invalid {key}")
    if data["orcid"] and not re.fullmatch(
        r"\d{4}-\d{4}-\d{4}-\d{3}[\dX]", str(data["orcid"])
    ):
        raise CollectionError("ORCID must use 0000-0000-0000-000X format")
    for key in ("keywordsZh", "keywordsEn"):
        if not isinstance(data[key], list) or not data[key]:
            raise CollectionError(f"{key} must contain at least one keyword")


def arguments() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, help="thesis root; defaults to script directory")
    parser.add_argument("--pdf", help="final PDF; defaults to main.pdf when present")
    parser.add_argument("--toc-file", help="generated .toc or UTF-8 plain-text contents")
    parser.add_argument("--references-file", help="UTF-8 text, one reference per line")
    parser.add_argument("--abstract-zh-file", help="UTF-8 Chinese abstract override")
    parser.add_argument("--abstract-en-file", help="UTF-8 English abstract override")
    parser.add_argument("--department", help="displayed department option in TDR")
    parser.add_argument("--graduation-year", help="ROC academic/graduation year")
    parser.add_argument("--publication-year", help="ROC publication year")
    parser.add_argument("--pages", help="last printed page number, not PDF page count")
    parser.add_argument("--note")
    parser.add_argument("--output", default=".tdr-upload.js")
    parser.add_argument("--non-interactive", action="store_true")
    return parser


def collect(args: argparse.Namespace) -> tuple[dict[str, object], Path]:
    root = (args.root or Path(__file__).resolve().parent).resolve()
    interactive = not args.non_interactive and sys.stdin.isatty()
    setup = parse_ntusetup(root / "ntusetup.tex")
    committee = parse_committee(root / "ntusetup.tex")
    check_academic_units(setup, root)
    abstract_zh, abstract_en = extract_abstracts(root / "front" / "abstract.tex")

    # The cover takes the advisor from \ntusetup, the TDR form from the
    # committee list, so both names are compared: correcting one spelling alone
    # is easy to do. The first entry is the 指導教授, which parse_committee has
    # already checked.
    for key, field in (("advisor", "nameZh"), ("advisor*", "nameEn")):
        if setup.get(key) and setup[key] != committee[0][field]:
            print(
                f"Warning: \\ntusetup {key} {setup[key]!r} differs from the "
                f"指導教授 in the committee list ({committee[0][field]!r})",
                file=sys.stderr,
            )
    if override := read_text(args.abstract_zh_file, root):
        abstract_zh = normalize_abstract(override)
    if override := read_text(args.abstract_en_file, root):
        abstract_en = normalize_abstract(override)

    pdf_path = Path(args.pdf) if args.pdf else root / "main.pdf"
    if not pdf_path.is_absolute():
        pdf_path = root / pdf_path
    if args.pdf and not pdf_path.is_file():
        raise CollectionError(f"PDF does not exist: {pdf_path}")
    final_pdf_text = pdf_text(pdf_path) if pdf_path.is_file() else ""

    if args.toc_file:
        toc_path = Path(args.toc_file)
        toc_path = toc_path if toc_path.is_absolute() else root / toc_path
    else:
        toc_path = root / "main.toc"
    if toc_path.is_file():
        toc = (
            extract_toc(toc_path)
            if toc_path.suffix == ".toc"
            else toc_path.read_text(encoding="utf-8")
        )
        toc = "\n".join(line.rstrip() for line in toc.splitlines() if line.strip())
    else:
        toc = ""

    references = (
        normalize_references(read_text(args.references_file, root))
        if args.references_file
        else references_from_pdf(final_pdf_text)
    )
    thesis_date = setup.get("date") or date.today().isoformat()
    try:
        graduation_year, semester, publication_year = academic_period(thesis_date)
    except CollectionError:
        raise
    except ValueError as error:
        raise CollectionError(f"Invalid thesis date: {thesis_date}") from error

    defense_date = setup.get("oral-date", "")

    department = DEPARTMENT_ALIASES.get(
        setup.get("institute", ""), setup.get("institute", "")
    )
    data: dict[str, object] = {
        "titleZh": setup.get("title", ""),
        "titleEn": setup.get("title*", ""),
        "department": choose(args.department, department, "Department", interactive),
        "graduationYear": choose(
            args.graduation_year, graduation_year, "Graduation year (ROC)", interactive
        ),
        "semester": semester,
        "publicationYear": choose(
            args.publication_year, publication_year, "Publication year (ROC)", interactive
        ),
        "authorZh": setup.get("author", ""),
        "authorEn": setup.get("author*", ""),
        # ORCID 與 email 來自 ntusetup.tex；沒填或還留著預設佔位字串時改為詢問。
        # Both come from ntusetup.tex; an empty or still-placeholder value falls
        # through to the same prompt every other uncertain field uses.
        "orcid": choose(
            None,
            placeholder(setup.get("ORCID"), *PLACEHOLDER_ORCID),
            "ORCID",
            interactive,
        ),
        "email": choose(
            None,
            placeholder(setup.get("email"), *PLACEHOLDER_EMAIL),
            "Email address",
            interactive,
        ),
        "email2": "",
        "committee": committee,
        "language": parse_language(root / "main.tex"),
        "defenseDate": defense_date,
        "pages": choose(
            args.pages,
            last_printed_page(final_pdf_text),
            "Last printed thesis page number",
            interactive,
        ),
        "keywordsZh": [x.strip() for x in setup.get("keywords", "").split(",") if x.strip()],
        "keywordsEn": [x.strip() for x in setup.get("keywords*", "").split(",") if x.strip()],
        "abstractZh": abstract_zh,
        "abstractEn": abstract_en,
        "tableOfContents": toc,
        "references": references,
        "note": args.note.strip() if args.note else "",
    }

    if not abstract_zh:
        raise CollectionError(
            "Chinese abstract is empty; complete front/abstract.tex or pass --abstract-zh-file"
        )
    if not abstract_en:
        raise CollectionError(
            "English abstract is empty; complete front/abstract.tex or pass --abstract-en-file"
        )
    if not toc:
        raise CollectionError("Build main.toc or pass --toc-file with the complete contents")
    if not references:
        raise CollectionError("Build/provide main.pdf or pass --references-file")
    validate(data)

    output = Path(args.output)
    output = output if output.is_absolute() else root / output
    require_ignored_output(output, root)
    return data, output


def render(data: dict[str, object], template: Path) -> str:
    source = template.read_text(encoding="utf-8")
    marker = "__TDR_DATA__"
    if source.count(marker) != 1:
        raise CollectionError(f"Expected one {marker} marker in {template}")
    encoded = json.dumps(data, ensure_ascii=False, indent=2)
    encoded = encoded.replace("\u2028", r"\u2028").replace("\u2029", r"\u2029")
    return source.replace(marker, encoded)


def main(argv: list[str] | None = None) -> int:
    try:
        data, output = collect(arguments().parse_args(argv))
        template = Path(__file__).resolve().with_name("tdr_upload_template.js")
        output.write_text(render(data, template), encoding="utf-8")
    except (CollectionError, FileNotFoundError) as error:
        print(f"error: {error}", file=sys.stderr)
        return 2
    print(f"Wrote copy-ready upload script: {output}")
    print("Review every field, then paste the entire file into the TDR DevTools console.")
    print("Paste it again on the committee page; it fills whichever page is open.")
    print("The generated script never clicks Save or Next.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
